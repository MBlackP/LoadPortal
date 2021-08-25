import calendar
from datetime import datetime
import getpass
from tkinter import *
from tkinter import filedialog

from openpyxl.styles import PatternFill, Alignment
from tkcalendar import DateEntry
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

'''
1. load excel file using current date/year aka Monthly_Invoice_year.xlsx = Monthly_Invoice_2021.xlsx
2. load month specified by user aka March 2021 will open March 2021 worksheet in 2021Invoice.xlsx.
3. fill the filedialog fields
    a) date picker (default is set to current date) for pickup and dropoff
    b) pickup and dropoff time = hours calculated by program
    c) ...
4. add the revenue calculator at a set location in worksheet
5. format everything except date to middle alignment
6. format revenue to dollar amounts
'''

root = Tk()
root.title('Load Portal')
root.geometry('915x750')
root.config(background = 'black')

currentDay = datetime.now().day
currentMonth = datetime.now().month
currentYear = datetime.now().year

def browseFile():
    global filePath
    filePath = filedialog.askopenfilename(initialdir = 'C:/Users/{}/', title = 'Select an Excel File')
    #change label contents
    fileSelectorLabel.configure(text='File Selected:\n' + filePath)
    global filePath2
    filePath2 = filePath.format(getpass.getuser())
    print(filePath2)

def newWorkbook():
    year = str(newWorkbookEF.get())
    newBook = Workbook()
    newSheet = newBook.active
    newSheet = newBook.create_sheet(f'January {year}', 0)
    del newBook['Sheet']

    greyFill = PatternFill(start_color='c4c4c4', end_color='c4c4c4', fill_type='solid')

    # title row
    newSheet.cell(row=2, column=1).value = 'Date'
    newSheet.cell(row=2, column=2).value = 'Pickup Time'
    newSheet.cell(row=2, column=3).value = 'Pickup Location'
    newSheet.cell(row=2, column=4).value = '# of Totes'
    newSheet.cell(row=2, column=5).value = '# of Skids'
    newSheet.cell(row=2, column=6).value = 'Dropoff Location'
    newSheet.cell(row=2, column=7).value = 'Dropoff Time'
    newSheet.cell(row=2, column=8).value = 'Dropoff T&T'
    newSheet.cell(row=2, column=9).value = 'Dropoff Bins'
    newSheet.cell(row=2, column=10).value = '# of Skids'
    newSheet.cell(row=2, column=11).value = 'Pickup Location'
    newSheet.cell(row=2, column=12).value = 'Pickup Full Bin'
    newSheet.cell(row=2, column=13).value = 'Pickup Skids'
    newSheet.cell(row=2, column=14).value = 'Dropoff'
    newSheet.cell(row=2, column=15).value = 'Contaminated Bins'
    newSheet.cell(row=2, column=16).value = 'Weight (kg)'
    newSheet.cell(row=2, column=17).value = 'Daily Hours'

    # color title row cells
    newSheet['A2'].fill = greyFill
    newSheet['B2'].fill = greyFill
    newSheet['C2'].fill = greyFill
    newSheet['D2'].fill = greyFill
    newSheet['E2'].fill = greyFill
    newSheet['F2'].fill = greyFill
    newSheet['G2'].fill = greyFill
    newSheet['H2'].fill = greyFill
    newSheet['I2'].fill = greyFill
    newSheet['J2'].fill = greyFill
    newSheet['K2'].fill = greyFill
    newSheet['L2'].fill = greyFill
    newSheet['M2'].fill = greyFill
    newSheet['N2'].fill = greyFill
    newSheet['O2'].fill = greyFill
    newSheet['P2'].fill = greyFill
    newSheet['Q2'].fill = greyFill

    redFill = PatternFill(start_color='db0000', end_color='db0000', fill_type='solid')
    greenFill = PatternFill(start_color='09D909', end_color='09D909', fill_type='solid')
    blueFill = PatternFill(start_color='2499FF', end_color='2499FF', fill_type='solid')
    pinkFill = PatternFill(start_color='FB80FF', end_color='FB80FF', fill_type='solid')

    # color cells before merging them
    newSheet['B1'].fill = redFill
    newSheet['C1'].fill = redFill
    newSheet['D1'].fill = redFill
    newSheet['E1'].fill = redFill

    newSheet['F1'].fill = greenFill
    newSheet['G1'].fill = greenFill
    newSheet['H1'].fill = greenFill
    newSheet['I1'].fill = greenFill
    newSheet['J1'].fill = greenFill

    newSheet['K1'].fill = blueFill
    newSheet['L1'].fill = blueFill
    newSheet['M1'].fill = blueFill

    newSheet['N1'].fill = pinkFill

    # merged cells row
    newSheet.merge_cells(start_row=1,start_column=2,end_row=1,end_column=5)
    newSheet.merge_cells(start_row=1,start_column=6,end_row=1,end_column=10)
    newSheet.merge_cells(start_row=1,start_column=11,end_row=1,end_column=13)

    newSheet['B1'].value = 'PICK UP'
    newSheet['F1'].value = 'DROP OFF'
    newSheet['K1'].value = 'PICK UP'
    newSheet['N1'].value = 'DROP OFF'

    newSheet['A86'] = 'Total Hours'
    newSheet['A87'] = 'Total Pay'
    newSheet['A88'] = 'GST'
    newSheet['A89'] = 'Net Income'

    newSheet['B86'] = '=SUM(Q3:Q85)'
    newSheet['B87'] = '=B86*58'
    newSheet['B88'] = '=B87*0.05'
    newSheet['B89'] = '=SUM(B87,B88)'

    # format cells to show as currency ($).
    newSheet['B87'].number_format = '$#,##0.00'
    newSheet['B88'].number_format = '$#,##0.00'
    newSheet['B89'].number_format = '$#,##0.00'

    # center alignment
    for i in range(1, 18):
        for j in range(1, 100):
            newSheet.cell(row=j, column=i).alignment = Alignment(horizontal='center')

    # resize width of columns
    for col in newSheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        newSheet.column_dimensions[column].width = max_length * 1.2

    filePathW = 'C:/Users/{}/' + f'Monthly_Invoice_{year}.xlsx'
    finalFilePathW = filePathW.format(getpass.getuser())
    newBook.save(filename=finalFilePathW)

    print('New Workbook Created.')

def newWorksheet():
    excelSheetName = str(newWorkSheetEF.get())
    year = str(newWorkSheetEF.get()).split(' ')[1]
    wb = xl.load_workbook(filePath2)
    newSheet = wb.create_sheet(f'{excelSheetName}')

    greyFill = PatternFill(start_color='c4c4c4', end_color='c4c4c4', fill_type='solid')

    # title row
    newSheet.cell(row=2, column=1).value = 'Date'
    newSheet.cell(row=2, column=2).value = 'Pickup Time'
    newSheet.cell(row=2, column=3).value = 'Pickup Location'
    newSheet.cell(row=2, column=4).value = '# of Totes'
    newSheet.cell(row=2, column=5).value = '# of Skids'
    newSheet.cell(row=2, column=6).value = 'Dropoff Location'
    newSheet.cell(row=2, column=7).value = 'Dropoff Time'
    newSheet.cell(row=2, column=8).value = 'Dropoff T&T'
    newSheet.cell(row=2, column=9).value = 'Dropoff Bins'
    newSheet.cell(row=2, column=10).value = '# of Skids'
    newSheet.cell(row=2, column=11).value = 'Pickup Location'
    newSheet.cell(row=2, column=12).value = 'Pickup Full Bin'
    newSheet.cell(row=2, column=13).value = 'Pickup Skids'
    newSheet.cell(row=2, column=14).value = 'Dropoff'
    newSheet.cell(row=2, column=15).value = 'Contaminated Bins'
    newSheet.cell(row=2, column=16).value = 'Weight (kg)'
    newSheet.cell(row=2, column=17).value = 'Daily Hours'

    # color title row cells
    newSheet['A2'].fill = greyFill
    newSheet['B2'].fill = greyFill
    newSheet['C2'].fill = greyFill
    newSheet['D2'].fill = greyFill
    newSheet['E2'].fill = greyFill
    newSheet['F2'].fill = greyFill
    newSheet['G2'].fill = greyFill
    newSheet['H2'].fill = greyFill
    newSheet['I2'].fill = greyFill
    newSheet['J2'].fill = greyFill
    newSheet['K2'].fill = greyFill
    newSheet['L2'].fill = greyFill
    newSheet['M2'].fill = greyFill
    newSheet['N2'].fill = greyFill
    newSheet['O2'].fill = greyFill
    newSheet['P2'].fill = greyFill
    newSheet['Q2'].fill = greyFill

    redFill = PatternFill(start_color='db0000', end_color='db0000', fill_type='solid')
    greenFill = PatternFill(start_color='09D909', end_color='09D909', fill_type='solid')
    blueFill = PatternFill(start_color='2499FF', end_color='2499FF', fill_type='solid')
    pinkFill = PatternFill(start_color='FB80FF', end_color='FB80FF', fill_type='solid')

    # color cells before merging them
    newSheet['B1'].fill = redFill
    newSheet['C1'].fill = redFill
    newSheet['D1'].fill = redFill
    newSheet['E1'].fill = redFill

    newSheet['F1'].fill = greenFill
    newSheet['G1'].fill = greenFill
    newSheet['H1'].fill = greenFill
    newSheet['I1'].fill = greenFill
    newSheet['J1'].fill = greenFill

    newSheet['K1'].fill = blueFill
    newSheet['L1'].fill = blueFill
    newSheet['M1'].fill = blueFill

    newSheet['N1'].fill = pinkFill

    # merged cells row
    newSheet.merge_cells(start_row=1,start_column=2,end_row=1,end_column=5)
    newSheet.merge_cells(start_row=1,start_column=6,end_row=1,end_column=10)
    newSheet.merge_cells(start_row=1,start_column=11,end_row=1,end_column=13)

    newSheet['B1'].value = 'PICK UP'
    newSheet['F1'].value = 'DROP OFF'
    newSheet['K1'].value = 'PICK UP'
    newSheet['N1'].value = 'DROP OFF'

    newSheet['A86'] = 'Total Hours'
    newSheet['A87'] = 'Total Pay'
    newSheet['A88'] = 'GST'
    newSheet['A89'] = 'Net Income'

    newSheet['B86'] = '=SUM(Q3:Q85)'
    newSheet['B87'] = '=B86*58'
    newSheet['B88'] = '=B87*0.05'
    newSheet['B89'] = '=SUM(B87,B88)'

    # format cells to show as currency ($).
    newSheet['B87'].number_format = '$#,##0.00'
    newSheet['B88'].number_format = '$#,##0.00'
    newSheet['B89'].number_format = '$#,##0.00'

    # center alignment
    for i in range(1,18):
        for j in range(1,100):
            newSheet.cell(row=j,column=i).alignment = Alignment(horizontal='center')

    # resize width of columns
    for col in newSheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        newSheet.column_dimensions[column].width = max_length * 1.2

    wb.save(filename=filePath2)

    print('New Worksheet Created.')

def addToSheet():
    # get month and year for searching for worksheet with 'month year' format.
    date = str(calendarPicker.get())
    month = date.split('/')[0]
    month = calendar.month_name[int(month)]
    year = date.split('/')[2]
    year = '20' + year

    excelSheetName = month + ' ' + year
    currentWorkbook = xl.load_workbook(filePath2)
    currentWorksheet = currentWorkbook[f'{excelSheetName}']

    # next step is to start writing where the row is empty.
    global row
    row = 3
    while row in range(3,85):
        if currentWorksheet.cell(row=row,column=1).value == None:
            break
        row += 1

    currentWorksheet.cell(row = row, column = 1).value = str(calendarPicker.get())
    currentWorksheet.cell(row = row, column = 2).value = str(pickupTimeEF.get())
    currentWorksheet.cell(row = row, column = 3).value = str(pickupLocationEF.get())
    currentWorksheet.cell(row = row, column = 4).value = str(numOfTotesEF.get())
    currentWorksheet.cell(row = row, column = 5).value = str(numOfSkidsPEF.get())
    currentWorksheet.cell(row = row, column = 6).value = str(dropoffLocationEF.get())
    currentWorksheet.cell(row = row, column = 7).value = str(dropoffTimeEF.get())
    currentWorksheet.cell(row = row, column = 8).value = str(dropoffTTEF.get())
    currentWorksheet.cell(row = row, column = 9).value = str(dropoffBinsEF.get())
    currentWorksheet.cell(row = row, column = 10).value = str(numOfSkidsDEF.get())
    currentWorksheet.cell(row = row, column = 11).value = str(pickupLocation2EF.get())
    currentWorksheet.cell(row = row, column = 12).value = str(pickupFullBinEF.get())
    currentWorksheet.cell(row = row, column = 13).value = str(pickupSkidsEF.get())
    currentWorksheet.cell(row = row, column = 14).value = str(dropoffL2EF.get())
    currentWorksheet.cell(row = row, column = 15).value = str(contaminatedBinsEF.get())
    currentWorksheet.cell(row = row, column = 16).value = str(weightEF.get())
    currentWorksheet.cell(row = row, column = 17).value = int(dailyHoursEF.get())

    currentWorkbook.save(filename=filePath2)

    print('Added to Worksheet.')

def clearFields():

    newWorkbookEF.delete(0,END)
    newWorkSheetEF.delete(0,END)
    calendarPicker.delete(0,END)
    pickupTimeEF.delete(0,END)
    pickupLocationEF.delete(0,END)
    numOfTotesEF.delete(0,END)
    numOfSkidsPEF.delete(0,END)
    dropoffTimeEF.delete(0,END)
    dropoffLocationEF.delete(0,END)
    dropoffTTEF.delete(0,END)
    dropoffBinsEF.delete(0,END)
    numOfSkidsDEF.delete(0,END)
    pickupLocation2EF.delete(0,END)
    pickupFullBinEF.delete(0,END)
    pickupSkidsEF.delete(0,END)
    dropoffL2EF.delete(0,END)
    contaminatedBinsEF.delete(0,END)
    weightEF.delete(0,END)
    dailyHoursEF.delete(0,END)

    print('Fields Cleared.')

# Select file from file path
fileSelectorLabel = Label(root, text = 'Select the Invoice File Below:', background = 'black', foreground = 'white', width=21)
fileSelectorLabel.grid(column = 1, row = 1, pady = 5)
fileSelector = Button(root, text = 'Select the Invoice File', command = browseFile)
fileSelector.grid(column = 1, row = 2, padx=50, pady = 10)

# Create new workbook
newWorkbookLabel = Label(root, text='Create New Workbook:\nEnter Year (e.g. 2021)',bg='black',fg='white')
newWorkbookLabel.grid(column=5,row=1,padx=5,pady=5)
newWorkbookField = StringVar()
newWorkbookEF = Entry(root, textvariable=newWorkbookField)
newWorkbookEF.grid(column=5,row=2,pady=10)
newWorkbookButton = Button(root, text='Create New Workbook', command=newWorkbook)
newWorkbookButton.grid(column=5,row=3,padx=50,pady=10)

# Create new worksheet
newWorkSheetLabel = Label(root, text='Create New Worksheet:\nEnter Month Year\n(e.g. March 2021)',bg='black',fg='white')
newWorkSheetLabel.grid(column=10,row=1,pady=10)
newWorkSheetField = StringVar()
newWorkSheetEF = Entry(root, textvariable=newWorkSheetField)
newWorkSheetEF.grid(column=10,row=2,pady=10)
newWorkSheetButton = Button(root, text='Create New Worksheet', command=newWorksheet)
newWorkSheetButton.grid(column=10,row=3,padx=50,pady=10)

# Date picker
datePickerLabel = Label(root, text = 'Choose load date below:', background = 'black', foreground= 'white')
datePickerLabel.grid(column = 1, row = 3, pady = 10)

currentDate = StringVar()
calendarPicker = DateEntry(root, width = 12, year = currentYear, month = currentMonth, day = currentDay, background = 'darkblue', foreground = 'white', borderwidth = 2, textvariable = currentDate)
calendarPicker.grid(column = 1, row = 4, padx=50,pady = 10)

# Pickup Time
pickupL = Label(root, text = 'PICKUP', background = 'black', foreground = 'white')
pickupL.grid(column = 1, row = 5, padx=50,pady = 10)
pickupTimeL = Label(root, text = 'Enter Pickup Time:', background = 'black', foreground = 'white')
pickupTimeL.grid(column = 1, row = 6, padx=50,pady = 10)
ptField = StringVar()
pickupTimeEF = Entry(root, textvariable = ptField)
pickupTimeEF.grid(column = 1, row = 7, padx=50,pady = 10)

# Pickup Location
pickupLocationL = Label(root, text = 'Enter Pickup Location:', background = 'black', foreground = 'white')
pickupLocationL.grid(column = 1, row = 8, padx=50,pady = 10)
plField = StringVar()
pickupLocationEF = Entry(root, textvariable = plField)
pickupLocationEF.grid(column = 1, row = 9, padx=50,pady = 10)

# Num. of Totes
numOfTotesL = Label(root, text = 'Enter # of Totes:', background = 'black', foreground = 'white')
numOfTotesL.grid(column = 1, row = 10, padx=50,pady = 10)
notField = StringVar()
numOfTotesEF = Entry(root, textvariable = notField)
numOfTotesEF.grid(column = 1, row = 11, padx=50,pady = 10)

# Num. of Skids
numOfSkidsPL = Label(root, text = 'Enter # of Skids:', background = 'black', foreground = 'white')
numOfSkidsPL.grid(column = 1, row = 12, padx=50,pady = 10)
nosPField = StringVar()
numOfSkidsPEF = Entry(root, textvariable = nosPField)
numOfSkidsPEF.grid(column = 1, row = 13, padx=50,pady = 10)

# Dropoff Time
dropoffL = Label(root, text = 'DROPOFF', background = 'black', foreground = 'white')
dropoffL.grid(column = 5, row = 5, padx=50,pady = 10)
dropoffTimeL = Label(root, text = 'Enter Dropoff Time:', background = 'black', foreground = 'white')
dropoffTimeL.grid(column = 5, row = 6, padx=50,pady = 10)
dtField = StringVar()
dropoffTimeEF = Entry(root, textvariable = dtField)
dropoffTimeEF.grid(column = 5, row = 7, padx=50,pady = 10)

# Dropoff Location
dropoffLocationL = Label(root, text = 'Enter Dropoff Location:', bg = 'black', fg = 'white')
dropoffLocationL.grid(column = 5, row = 8, padx=50,pady = 10)
dllField = StringVar()
dropoffLocationEF = Entry(root, textvariable = dllField)
dropoffLocationEF.grid(column = 5, row = 9, padx=50,pady = 10)

# Dropoff T&T
dropoffTTL = Label(root, text = 'Enter Dropoff at T&T:', bg = 'black', fg = 'white')
dropoffTTL.grid(column = 5, row = 10, padx=50,pady = 10)
dropoffTTField = StringVar()
dropoffTTEF = Entry(root, textvariable = dropoffTTField)
dropoffTTEF.grid(column = 5, row = 11, padx=50,pady = 10)

# Dropoff bins
dropoffBinsL = Label(root, text = 'Enter # of Dropoff Bins:', bg = 'black', fg = 'white')
dropoffBinsL.grid(column = 5, row = 12, padx=50,pady = 10)
dropoffBinsField = StringVar()
dropoffBinsEF = Entry(root, textvariable = dropoffBinsField)
dropoffBinsEF.grid(column = 5, row = 13, padx=50,pady = 10)

# Num. of skids
numOfSkidsDL = Label(root, text = 'Enter # of Dropoff Skids:', bg = 'black', fg = 'white')
numOfSkidsDL.grid(column = 5, row = 14, padx=50,pady = 10)
nosDField = StringVar()
numOfSkidsDEF = Entry(root, textvariable = nosDField)
numOfSkidsDEF.grid(column = 5, row = 15, padx=50,pady = 10)

# Pickup Location
pickupL2 = Label(root, text = 'PICKUP', background = 'black', foreground = 'white')
pickupL2.grid(column = 10, row = 5, padx=50,pady = 10)
pickupLocation2L = Label(root, text = 'Pickup Location:', bg = 'black', fg = 'white')
pickupLocation2L.grid(column = 10, row = 6, padx=50,pady = 10)
pickupL2Field = StringVar()
pickupLocation2EF = Entry(root, textvariable = pickupL2Field)
pickupLocation2EF.grid(column = 10, row = 7, padx=50, pady = 10)

# Pickup full bin
pickupFullBinL = Label(root, text = 'Pickup Full Bin:', bg = 'black', fg = 'white')
pickupFullBinL.grid(column = 10, row = 8,padx=50, pady = 10)
pickupFullBinField = StringVar()
pickupFullBinEF = Entry(root, textvariable = pickupFullBinField)
pickupFullBinEF.grid(column = 10, row = 9, padx=50, pady = 10)

# Pickup skids
pickupSkidsL = Label(root, text = 'Pickup Skids:', bg = 'black', fg = 'white')
pickupSkidsL.grid(column = 10, row = 10, padx=50,pady = 10)
pickupSkidsField = StringVar()
pickupSkidsEF = Entry(root, textvariable = pickupSkidsField)
pickupSkidsEF.grid(column = 10, row = 11, padx=50,pady = 10)

# Dropoff
dropoffBigLabel = Label(root, text = 'DROPOFF', bg = 'black', fg = 'white')
dropoffBigLabel.grid(column = 20, row = 5, padx=50,pady = 10)
dropoffL2 = Label(root, text = 'Dropoff:', bg = 'black', fg = 'white')
dropoffL2.grid(column = 20, row = 6, padx=50,pady = 10)
dropoffL2Field = StringVar()
dropoffL2EF = Entry(root, textvariable = dropoffL2Field)
dropoffL2EF.grid(column = 20, row = 7, padx=50,pady = 10)

# Contaminated bins
contaminatedBinsL = Label(root, text = 'Contaminated Bins:', bg = 'black', fg = 'white')
contaminatedBinsL.grid(column = 20, row = 8, padx=50,pady = 10)
contaminatedBinsField = StringVar()
contaminatedBinsEF = Entry(root, textvariable = contaminatedBinsField)
contaminatedBinsEF.grid(column = 20, row = 9, padx=50,pady = 10)

# Weight (kg)
weightL = Label(root, text = 'Weight (kg):', bg = 'black', fg = 'white')
weightL.grid(column = 20, row = 10, padx=50,pady = 10)
weightField = StringVar()
weightEF = Entry(root, textvariable = weightField)
weightEF.grid(column = 20, row = 11, padx=50,pady = 10)

# daily hours - calculated by using Pickup Time and Dropoff Time
dailyHoursL = Label(root, text = 'Daily Hours:', bg = 'black', fg='white')
dailyHoursL.grid(column = 20, row = 12, padx=50,pady = 10)
dhField = StringVar()
dailyHoursEF = Entry(root, textvariable = dhField)
dailyHoursEF.grid(column = 20, row = 13, padx=50,pady = 10)

submitButton = Button(root, text = 'Add Load to Sheet', bg = 'light green', width = 15, height = 3, command = addToSheet)
submitButton.grid(column = 20, row = 15)

clearButton = Button(root, text='Clear Fields',width=15,height=3,bg='yellow', command = clearFields)
clearButton.grid(column=20,row=1,pady=10)

exit_button = Button(root, text = 'Close', background = 'red', width = 15, height = 3, command = root.destroy)
exit_button.grid(column = 20, row = 17, pady = 10)

root.mainloop()
